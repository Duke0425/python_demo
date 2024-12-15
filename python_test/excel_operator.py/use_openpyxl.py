import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


CommonSheetNameList = [
    "第1筒次010段_12345",
    "第2筒次012段_12345",
    "第3筒次011段_01120",
]

ExcelPath = 'C:\\Users\\Admin\\Desktop\\tesss1\\openpyxl_simple.xlsx'

ProjectInfoDict = {
    "项目名称": "四川省广汉市中石油德页1井",
    "地理位置": "四川省广汉市",
    "井口名称(井号)": "0011",
    "取心筒次": "2021024",
    "岩心直径": 26,
    "其他": ""
}

RockInfoDict = {
    "岩心编号": "202410230851192",
    "岩心名称": "第一筒次010段",
    "岩心块号": "1-10",
    "序列名称": "cpmg_1",
    "测量长度(m)": "-",
    "测量顶深(m)":"-",
    "测量底深(m)":"-",
    "测量时间":"-",
    "数据名称":"-",
    "回波间隔":"-",
    "回波个数":"-",
    "TW":"-",
    "平均次数":"-",
    "其他":"-",
}

class OpenPyExcelWritter:

    def __init__(self, aExcelPath:str, aSheetName) -> None:
        self.myWorkbook = Workbook()
        self.myWorksheet = self.myWorkbook.create_sheet(title=aSheetName, index=0)
        self.myWorkbook.create_sheet(title="Test1")
        self.myWorkbook.create_sheet(title="aSheetName")
        self.myWorkbook.create_sheet(title="nothing")
        self.myWorkbookPath = aExcelPath
        self.initFormatSettings()

    def initFormatSettings(self):
        self.myTitleFont = Font(bold=True, size=20)
        self.myAligment = Alignment(horizontal="left")

    def origanizedData(self):
        self.myWorksheet.merge_cells("A1:B1")
        A1Cell = self.myWorksheet["A1"]
        A1Cell.value = "项目数据信息表"
        A1Cell.font = self.myTitleFont
        A1Cell.alignment = self.myAligment

        self.myWorksheet.column_dimensions['A'].width = 20
        self.myWorksheet.column_dimensions['B'].width = 20

        self.__addProjectInfoToSheet()
        self.__addRockInfoToSheet()

    def save(self):
        self.myWorkbook.save(self.myWorkbookPath)

    def __addProjectInfoToSheet(self):
        for key, value in ProjectInfoDict.items():
            self.myWorksheet.append([key,value])

    def __addRockInfoToSheet(self):
        titleRow = self.myWorksheet.max_row + 1
        self.myWorksheet.row_dimensions[titleRow+1].font = Font(bold=True)
        self.myWorksheet.append(list(RockInfoDict.keys()))
        self.myWorksheet.insert_rows(titleRow)

        while self.myWorksheet.max_row < 10000:
            self.myWorksheet.append(list(RockInfoDict.values()))


def main():
    start_time = time.time()
    excelWriter = OpenPyExcelWritter(ExcelPath, CommonSheetNameList[0])
    excelWriter.origanizedData()
    excelWriter.save()
    print("--- %s seconds ---" % (time.time() - start_time))


if __name__ == '__main__':
    main()